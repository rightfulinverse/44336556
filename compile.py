import os
import json
from pathlib import Path
from openpyxl import load_workbook

def read_cfr_json(directory):
    cfr_json = {}
    for filename in os.listdir(directory):
        if filename.endswith(".json"):
            file_path = os.path.join(directory, filename)
            with open(file_path, "r") as file:
                cfr_json[filename] = json.load(file)
    return cfr_json

def read_cfr_corrections_json(directory):
    cfr_corrections_json = {}
    for filename in os.listdir(directory):
        if filename.endswith(".json"):
            file_path = os.path.join(directory, filename)
            with open(file_path, "r") as file:
                cfr_corrections_json[filename] = json.load(file)
    return cfr_corrections_json

def clean_files(directory):
    for root, dirs, files in os.walk(directory):
        for file in files:
            if file.endswith(".html"):
                file_path = os.path.join(root, file)
                try:
                    os.remove(file_path)
                    print(f"Deleted: {file_path}")
                except Exception as e:
                    print(f"Error deleting {file_path}: {e}")

# Map of title number to percentage of CFR total content
def generate_title_breakdown(cfr_json):
    total = 0
    title_breakdown = {}

    # Get the total size    
    for cfr_json_file_name, cfr_json_content in cfr_json.items():
        if cfr_json_file_name == "agencies.json":
            continue
        total += cfr_json_content["size"]

    # Get each title size
    for cfr_json_file_name, cfr_json_content in cfr_json.items():
        if cfr_json_file_name == "agencies.json":
            continue
            
        title_pct = (float(cfr_json_content["size"]) / total)
        title_breakdown[cfr_json_content["identifier"]] = title_pct
    title_breakdown_sorted = dict(sorted(title_breakdown.items(), key=lambda item: item[1], reverse=True))
    return title_breakdown_sorted

# Map of title number to map where the key is the year and the value is the number of corrections made that year
def generate_corrections_breakdown(cfr_corrections_json):
    corrections_breakdown = {}
    for cfr_corrections_json_file_name, cfr_corrections_json_content in cfr_corrections_json.items():
        for cfr_correction in cfr_corrections_json_content["ecfr_corrections"]:
            # Create new map for years
            if not cfr_correction["title"] in corrections_breakdown:
                corrections_breakdown[cfr_correction["title"]] = {}

            if cfr_correction["year"] in corrections_breakdown[cfr_correction["title"]]:
                corrections_breakdown[cfr_correction["title"]][cfr_correction["year"]] = corrections_breakdown[cfr_correction["title"]][cfr_correction["year"]] + 1
            else:
                corrections_breakdown[cfr_correction["title"]][cfr_correction["year"]] = 1
    return corrections_breakdown

# Map of agency name to percentage of title content
def generate_agency_breakdown(cfr_title_json, agencies_json):
    agency_breakdown = {}
    for child in cfr_title_json["children"]:
        for agency in agencies_json["agencies"]:
            if agency["name"] in child["label"]:
                agency_pct = (float(child["size"]) / float(cfr_title_json["size"]))
                if agency["name"] in agency_breakdown:
                    agency_breakdown[agency["name"]] += agency_pct
                else:
                    agency_breakdown[agency["name"]] = agency_pct
                
    agency_breakdown_sorted = dict(sorted(agency_breakdown.items(), key=lambda item: item[1], reverse=True))
    return agency_breakdown_sorted

# https://www.eeoc.gov/federal-sector/reports/annual-reports-federal-workforce-including-data-tables
def generate_workforce_breakdown():
    workforce_breakdown = {}
    workbook = load_workbook(filename="FY 2021 Annual Report Workforce Tables 2023Dec12.xlsx")

    # Generated from table of contents
    sheet_numbers = ["A-1b", "A-1d"]

    for sheet_number in sheet_numbers:
        sheet = workbook[sheet_number]

        row = 4
        while row < 200:
            agency_name = sheet.cell(row=row, column=1).value
            agency_code = sheet.cell(row=row, column=2).value
            total_workforce = sheet.cell(row=row, column=3).value
            male = sheet.cell(row=row, column=4).value
            female = sheet.cell(row=row, column=5).value
            latin_male = sheet.cell(row=row, column=6).value
            latin_female = sheet.cell(row=row, column=7).value
            white_male = sheet.cell(row=row, column=8).value
            white_female = sheet.cell(row=row, column=9).value
            black_male = sheet.cell(row=row, column=10).value
            black_female = sheet.cell(row=row, column=11).value
            asian_male = sheet.cell(row=row, column=12).value
            asian_female = sheet.cell(row=row, column=13).value
            other_male = sheet.cell(row=row, column=14).value
            other_female = sheet.cell(row=row, column=15).value
            indian_male = sheet.cell(row=row, column=16).value
            indian_female = sheet.cell(row=row, column=17).value

            if agency_code is not None:
                agency_name = agency_name.replace(": Department-Wide Data", "") # some normalization
                workforce_breakdown[agency_name] = {
                    "agency_code": agency_code,
                    "total_workforce": total_workforce,
                    "male": male,
                    "female": female,
                    "latin_male": latin_male,
                    "latin_female": latin_female,
                    "white_male": white_male,
                    "white_female": white_female,
                    "black_male": black_male,
                    "black_female": black_female,
                    "asian_male": asian_male,
                    "asian_female": asian_female,
                    "other_male": other_male,
                    "other_female": other_female,
                    "indian_male": indian_male,
                    "indian_female": indian_female
                }
            
            row += 1

    return workforce_breakdown

def process_files(source_dir, target_dir, cfr_json, cfr_corrections_json):
    os.makedirs(target_dir, exist_ok=True)

    for root, _, files in os.walk(source_dir):
        for file in files:
            source_path = Path(root) / file
            relative_path = source_path.relative_to(source_dir)
            target_path = Path(target_dir) / relative_path

            os.makedirs(target_path.parent, exist_ok=True)

            with open(source_path, "r", encoding="utf-8") as source_file:
                content = source_file.read()

            modified_content = content

            # TITLE BREAKDOWN
            title_breakdown = generate_title_breakdown(cfr_json)
            title_breakdown_json = json.dumps(title_breakdown, indent=4)
            modified_content = modified_content.replace("$CFR-BREAKDOWN-JSON$", f"<script>var title_breakdown_json = {title_breakdown_json};</script>")

            # CORRECTIONS BREAKDOWN
            corrections_breakdown = generate_corrections_breakdown(cfr_corrections_json)
            corrections_breakdown_json = json.dumps(corrections_breakdown, indent=4)
            modified_content = modified_content.replace("$CORRECTIONS-BREAKDOWN-JSON$", f"<script>var corrections_breakdown_json = {corrections_breakdown_json};</script>")

            # WORKFORCE BREAKDOWN
            workforce_breakdown = generate_workforce_breakdown()
            workforce_breakdown_json = json.dumps(workforce_breakdown, indent=4)
            modified_content = modified_content.replace("$WORKFORCE-BREAKDOWN-JSON$", f"<script>var workforce_breakdown_json = {workforce_breakdown_json};</script>")

            # AGENCY WITHIN TITLE BREAKDOWN
            for cfr_json_file_name, cfr_json_content in cfr_json.items():
                if cfr_json_file_name == "agencies.json":
                    continue
                
                # Checking each title
                identifier = cfr_json_content["identifier"]
                label = "<b>" + cfr_json_content["label"] + "</b>"
                agency_breakdown_content = ""

                agency_breakdown = generate_agency_breakdown(cfr_json_content, cfr_json["agencies.json"])
                for agency in agency_breakdown:
                    agency_pct = agency_breakdown[agency] * 100
                    agency_breakdown_content += agency + " <small class=\"text-muted\">" + f"{int(agency_pct * 1000) / 1000:.3f}% makeup</small>" + "<br>"

                modified_content = modified_content.replace(f"$CFR-BREAKDOWN-TITLE{identifier}$", label + "<div class=\"container\">" + agency_breakdown_content + "</div>")

            with open(target_path, "w", encoding="utf-8") as target_file:
                target_file.write(modified_content)

source_directory = "src_html"
target_directory = "public_html"

cfr_json = read_cfr_json(target_directory + "/cfr_json")
cfr_corrections_json = read_cfr_json(target_directory + "/cfr_corrections_json")
clean_files(target_directory)
process_files(source_directory, target_directory, cfr_json, cfr_corrections_json)
